#!/usr/bin/env python3
"""Ingest the consolidated Bahrain workbook into data/documents JSON records.

Reads sheet "Bahrain (2)" (already stamped with regulation_id / is_primary by
add-regulation-ids.py), groups rows by regulation, writes one record per
regulation, and upserts any missing Bahrain authorities. Skips a regulation
when its generated id already exists as a curated file.

Usage: python scripts/ingest-bahrain.py [WORKBOOK.xlsx]
"""
from __future__ import annotations

import json
import re
import shutil
import sys
import tempfile
from collections import OrderedDict, Counter
from pathlib import Path

import openpyxl

sys.path.insert(0, str(Path(__file__).resolve().parent))
from bahrain_ingest_lib import slugify, resolve_authority, resolve_leaf, build_record, _norm

REPO = Path(__file__).resolve().parent.parent
DOCS_DIR = REPO / "data" / "documents"
AUTH_PATH = REPO / "data" / "authorities.json"
TAX_PATH = REPO / "data" / "taxonomy.json"
SHEET = "Bahrain (2)"
HEADER_ROW = 1  # zero based index into iter_rows
DEFAULT_WB = Path.home() / "OneDrive" / "Desktop" / "Bahrain_Regulatory_Database_deep_crawl_v2_with_ids.xlsx"

# Workbook regulation_id -> existing curated record id. These are already in the
# repo as hand built rich records, so the workbook duplicate is skipped. The PDPL
# slugifies to a different id than the curated bh-pdpl-2018, so it must be matched
# by regulation_id here, not by generated id collision.
CURATED_OVERLAP = {"BH-R001": "bh-pdpl-2018"}

# Authorities the script ensures exist in authorities.json. Ids match AUTHORITY_ALIASES.
AUTHORITY_ENTRIES = {
    "bh-tra": {"name": "Telecommunications Regulatory Authority", "type": "sectoral_regulator",
               "mandate": "Regulates the telecommunications sector, spectrum, and related digital infrastructure in Bahrain.",
               "website": "https://www.tra.org.bh"},
    "bh-ncsc": {"name": "National Cyber Security Center", "type": "sectoral_regulator",
                "mandate": "Leads national cybersecurity strategy, controls, and incident response.",
                "website": "https://www.ncsc.gov.bh"},
    "bh-moj": {"name": "Ministry of Justice, Islamic Affairs and Waqf", "type": "ministry",
               "mandate": "Administers justice and, for the data protection law, the supervisory authority function.",
               "website": "https://www.moj.gov.bh"},
    "bh-iga": {"name": "Information and eGovernment Authority", "type": "government_agency",
               "mandate": "Runs national e-government services, digital identity, open data, and information policy.",
               "website": "https://www.iga.gov.bh"},
    "bh-nhra": {"name": "National Health Regulatory Authority", "type": "sectoral_regulator",
                "mandate": "Regulates healthcare providers and services, including digital and tele-health.",
                "website": "https://www.nhra.bh"},
    "bh-moic": {"name": "Ministry of Industry and Commerce", "type": "ministry",
                "mandate": "Oversees commerce, e-commerce, and consumer protection.",
                "website": "https://www.moic.gov.bh"},
    "bh-pm": {"name": "Prime Minister", "type": "head_of_government",
              "mandate": "Issues edicts and decisions of the Council of Ministers.",
              "website": None},
    "bh-nbr": {"name": "National Bureau for Revenue", "type": "government_agency",
               "mandate": "Administers VAT and excise, including e-invoicing and digital tax administration.",
               "website": "https://www.nbr.gov.bh"},
    "bh-shura": {"name": "Shura Council", "type": "legislature",
                 "mandate": "Upper chamber of the National Assembly; participates in enacting legislation.",
                 "website": "https://www.shura.bh"},
    "bh-caa": {"name": "Civil Aviation Affairs", "type": "sectoral_regulator",
               "mandate": "Regulates civil aviation, including drones and unmanned systems.",
               "website": "https://www.caa.gov.bh"},
    "bh-scict": {"name": "Supreme Council for Information and Communication Technology", "type": "council",
                 "mandate": "Sets national information and communication technology policy direction.",
                 "website": None},
    "bh-edb": {"name": "Bahrain Economic Development Board", "type": "government_agency",
               "mandate": "Promotes investment and economic development, including digital and emerging technology initiatives.",
               "website": "https://www.bahrainedb.com"},
    "bh-gov": {"name": "Government of Bahrain", "type": "government",
               "mandate": "General government of the Kingdom of Bahrain; used where the specific issuing body is unconfirmed.",
               "website": "https://www.bahrain.bh"},
    "bh-mtt": {"name": "Ministry of Transportation and Telecommunications", "type": "ministry",
               "mandate": "Oversees transportation and telecommunications policy and ministerial decisions.",
               "website": "https://www.mtt.gov.bh"},
    "bh-mofne": {"name": "Ministry of Finance and National Economy", "type": "ministry",
                 "mandate": "Oversees public finance and national economic policy.",
                 "website": "https://www.mof.gov.bh"},
    "bh-moh": {"name": "Ministry of Health", "type": "ministry",
               "mandate": "Oversees public health policy and health services.",
               "website": "https://www.moh.gov.bh"},
}


def load_workbook_rows(path):
    with tempfile.TemporaryDirectory() as td:
        tmp = Path(td) / path.name
        shutil.copy2(path, tmp)
        wb = openpyxl.load_workbook(tmp, read_only=True, data_only=True)
        ws = wb[SHEET]
        rows = list(ws.iter_rows(values_only=True))
        wb.close()
    header = [(h or "").strip() if isinstance(h, str) else h for h in rows[HEADER_ROW]]
    out = []
    for r in rows[HEADER_ROW + 1:]:
        if all(c is None for c in r):
            continue
        out.append(dict(zip(header, r)))
    return out


def build_taxonomy_lookups():
    tax = json.loads(TAX_PATH.read_text(encoding="utf-8"))
    leaf_lookup = {}
    branch_label = {}
    for branch in tax["themes"]:
        branch_label[branch["id"]] = branch["label"]
        for leaf in branch.get("children", []):
            leaf_lookup[(_norm(branch["label"]), _norm(leaf["label"]))] = leaf["id"]
    binding_by_class = {c["id"]: bool(c.get("binding")) for c in tax["instrument_classes"]}
    return leaf_lookup, branch_label, binding_by_class


def upsert_authorities(used_ids):
    authorities = json.loads(AUTH_PATH.read_text(encoding="utf-8"))
    existing = {a["id"] for a in authorities}
    added = []
    for aid in sorted(used_ids):
        if aid in existing:
            continue
        entry = AUTHORITY_ENTRIES.get(aid)
        if not entry:
            continue
        authorities.append({
            "id": aid, "name": entry["name"], "name_ar": None,
            "jurisdiction": "BH", "type": entry["type"],
            "mandate": entry["mandate"], "website": entry["website"],
        })
        added.append(aid)
    if added:
        AUTH_PATH.write_text(json.dumps(authorities, ensure_ascii=False, indent=2) + "\n", encoding="utf-8")
    return added


def allocate_id(title, taken):
    base = "bh-" + slugify(title)
    if base == "bh-":
        base = "bh-instrument"
    cid = base
    n = 2
    while cid in taken:
        cid = f"{base}-{n}"
        n += 1
    taken.add(cid)
    return cid


def main():
    wb_path = Path(sys.argv[1]) if len(sys.argv) > 1 else DEFAULT_WB
    if not wb_path.exists():
        print(f"ERROR: workbook not found: {wb_path}", file=sys.stderr)
        return 1

    rows = load_workbook_rows(wb_path)
    leaf_lookup, branch_label, binding_by_class = build_taxonomy_lookups()

    groups = OrderedDict()
    for r in rows:
        groups.setdefault(r.get("regulation_id"), []).append(r)

    # Idempotent: remove previously generated Bahrain catalogue files so reruns do
    # not accumulate numeric suffixes. Curated Bahrain records are protected.
    protected = set(CURATED_OVERLAP.values())
    for p in DOCS_DIR.glob("bh-*.json"):
        if p.stem not in protected:
            p.unlink()

    existing_files = {p.stem for p in DOCS_DIR.glob("*.json")}
    taken = set(existing_files)

    written = 0
    skipped_existing = []
    unmapped_authorities = Counter()
    unmapped_leaves = Counter()
    used_authority_ids = set()
    records = []

    for rid, grp in groups.items():
        if rid in CURATED_OVERLAP:
            skipped_existing.append((rid, CURATED_OVERLAP[rid]))
            continue
        primary = next((r for r in grp if r.get("is_primary") == 1), grp[0])
        title = str(primary.get("Instrument name (EN)") or "").strip()

        # Resolve authority and report misses.
        aid, _ = resolve_authority(primary.get("Issuing authority"))
        if aid is None:
            unmapped_authorities[str(primary.get("Issuing authority"))] += 1
        else:
            used_authority_ids.add(aid)

        # Report any leaf that fails to map.
        for r in grp:
            if resolve_leaf(r.get("Category"), r.get("Subcategory"), leaf_lookup) is None:
                unmapped_leaves[(str(r.get("Category")), str(r.get("Subcategory")))] += 1

        doc_id = allocate_id(title, taken)
        rec = build_record(grp, doc_id=doc_id, leaf_lookup=leaf_lookup,
                           branch_label=branch_label, binding_by_class=binding_by_class)
        records.append(rec)

    # Add authorities before writing records so referential integrity holds.
    added = upsert_authorities(used_authority_ids)

    for rec in records:
        (DOCS_DIR / f"{rec['id']}.json").write_text(
            json.dumps(rec, ensure_ascii=False, indent=2) + "\n", encoding="utf-8")
        written += 1

    print(f"Workbook : {wb_path}")
    print(f"Regulations in sheet : {len(groups)}")
    print(f"Records written      : {written}")
    print(f"Authorities added    : {len(added)} {added}")
    print(f"Skipped (id already exists): {len(skipped_existing)} {skipped_existing}")
    if unmapped_authorities:
        print("\nUNMAPPED authorities (add to AUTHORITY_ALIASES/ENTRIES):")
        for name, c in unmapped_authorities.most_common():
            print(f"  {c:>3}  {name}")
    if unmapped_leaves:
        print("\nUNMAPPED (Category, Subcategory) pairs (check against taxonomy labels):")
        for pair, c in unmapped_leaves.most_common():
            print(f"  {c:>3}  {pair}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
